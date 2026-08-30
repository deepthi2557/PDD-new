import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .config import EXCEL_REPORT_PATH

def generate_excel_analysis_report(test_results, output_path=EXCEL_REPORT_PATH):
    """
    Generates a multi-sheet, styled Excel Analysis Report from test execution results.
    """
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # STYLES DEFINITIONS
    # -------------------------------------------------------------
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    title_font = Font(name="Calibri", size=16, bold=True, color="0F172A")
    subtitle_font = Font(name="Calibri", size=11, italic=True, color="475569")
    section_font = Font(name="Calibri", size=13, bold=True, color="1E293B")
    
    kpi_title_font = Font(name="Calibri", size=9, bold=True, color="64748B")
    kpi_value_font = Font(name="Calibri", size=18, bold=True, color="0F172A")
    
    pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Light Green
    pass_font = Font(name="Calibri", size=10, bold=True, color="15803D")
    
    fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Light Red
    fail_font = Font(name="Calibri", size=10, bold=True, color="B91C1C")
    
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    thick_bottom = Border(bottom=Side(style='medium', color='0F172A'))
    
    # -------------------------------------------------------------
    # SHEET 1: EXECUTIVE SUMMARY
    # -------------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_summary["A1"] = "Web Application Selenium Test Analysis Report"
    ws_summary["A1"].font = title_font
    ws_summary["A2"] = "End-to-End Automated UI & Multi-Tab Workflow Test Analysis Suite"
    ws_summary["A2"].font = subtitle_font
    
    total_tests = len(test_results)
    passed_tests = sum(1 for r in test_results if r["Status"] == "PASS")
    failed_tests = sum(1 for r in test_results if r["Status"] == "FAIL")
    pass_rate = round((passed_tests / total_tests * 100), 1) if total_tests > 0 else 0.0
    total_duration = round(sum(r["Duration_Sec"] for r in test_results), 2)
    multi_tab_count = sum(1 for r in test_results if r.get("Is_Multi_Tab") == "YES")
    
    # KPI Cards Block (Rows 4 - 6)
    kpis = [
        ("TOTAL TEST CASES", total_tests, "B4:C5"),
        ("PASSED TESTS", passed_tests, "E4:F5"),
        ("FAILED TESTS", failed_tests, "H4:I5"),
        ("PASS RATE", f"{pass_rate}%", "K4:L5"),
        ("MULTI-TAB TESTS", multi_tab_count, "N4:O5")
    ]
    
    for title, val, cell_range in kpis:
        top_left = cell_range.split(":")[0]
        col_letter = top_left[0]
        row_num = int(top_left[1])
        
        ws_summary[f"{col_letter}{row_num}"] = title
        ws_summary[f"{col_letter}{row_num}"].font = kpi_title_font
        
        val_cell = f"{col_letter}{row_num + 1}"
        ws_summary[val_cell] = val
        ws_summary[val_cell].font = kpi_value_font
        if title == "PASSED TESTS":
            ws_summary[val_cell].font = Font(name="Calibri", size=18, bold=True, color="16A34A")
        elif title == "FAILED TESTS" and failed_tests > 0:
            ws_summary[val_cell].font = Font(name="Calibri", size=18, bold=True, color="DC2626")

    # Section Header
    ws_summary["A8"] = "Category Breakdown & Execution Summary"
    ws_summary["A8"].font = section_font
    
    headers_summary = ["Category", "Total Test Cases", "Passed", "Failed", "Pass Rate (%)", "Avg Duration (s)"]
    for col_idx, h_text in enumerate(headers_summary, start=1):
        cell = ws_summary.cell(row=9, column=col_idx, value=h_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Group results by Category
    categories = {}
    for r in test_results:
        cat = r["Category"]
        if cat not in categories:
            categories[cat] = {"total": 0, "passed": 0, "failed": 0, "durations": []}
        categories[cat]["total"] += 1
        if r["Status"] == "PASS":
            categories[cat]["passed"] += 1
        else:
            categories[cat]["failed"] += 1
        categories[cat]["durations"].append(r["Duration_Sec"])
        
    row_curr = 10
    for cat, data in categories.items():
        c_tot = data["total"]
        c_pass = data["passed"]
        c_fail = data["failed"]
        c_rate = round((c_pass / c_tot * 100), 1) if c_tot > 0 else 0.0
        c_avg_dur = round(sum(data["durations"]) / c_tot, 3) if c_tot > 0 else 0.0
        
        ws_summary.cell(row=row_curr, column=1, value=cat).alignment = Alignment(horizontal="left")
        ws_summary.cell(row=row_curr, column=2, value=c_tot).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=row_curr, column=3, value=c_pass).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=row_curr, column=4, value=c_fail).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=row_curr, column=5, value=f"{c_rate}%").alignment = Alignment(horizontal="center")
        ws_summary.cell(row=row_curr, column=6, value=c_avg_dur).alignment = Alignment(horizontal="center")
        
        for c in range(1, 7):
            cell = ws_summary.cell(row=row_curr, column=c)
            cell.border = thin_border
            if row_curr % 2 == 0:
                cell.fill = zebra_fill
        row_curr += 1
        
    # -------------------------------------------------------------
    # SHEET 2: DETAILED TEST RESULTS
    # -------------------------------------------------------------
    ws_details = wb.create_sheet(title="Detailed Test Results")
    ws_details.views.sheetView[0].showGridLines = True
    
    headers_details = [
        "Test Case ID", "Category", "Feature Name", "Button / Target Element", 
        "Action Tested", "Status", "Duration (s)", "Multi-Tab Verified", "Error Details"
    ]
    
    ws_details.row_dimensions[1].height = 25
    for col_idx, h_text in enumerate(headers_details, start=1):
        cell = ws_details.cell(row=1, column=col_idx, value=h_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for row_idx, r in enumerate(test_results, start=2):
        ws_details.cell(row=row_idx, column=1, value=r["Test_ID"]).alignment = Alignment(horizontal="center")
        ws_details.cell(row=row_idx, column=2, value=r["Category"]).alignment = Alignment(horizontal="left")
        ws_details.cell(row=row_idx, column=3, value=r["Feature_Name"]).alignment = Alignment(horizontal="left")
        ws_details.cell(row=row_idx, column=4, value=r["Button_Target"]).alignment = Alignment(horizontal="left")
        ws_details.cell(row=row_idx, column=5, value=r["Action_Tested"]).alignment = Alignment(horizontal="left")
        
        status_cell = ws_details.cell(row=row_idx, column=6, value=r["Status"])
        status_cell.alignment = Alignment(horizontal="center")
        if r["Status"] == "PASS":
            status_cell.fill = pass_fill
            status_cell.font = pass_font
        else:
            status_cell.fill = fail_fill
            status_cell.font = fail_font
            
        ws_details.cell(row=row_idx, column=7, value=r["Duration_Sec"]).alignment = Alignment(horizontal="center")
        ws_details.cell(row=row_idx, column=8, value=r.get("Is_Multi_Tab", "NO")).alignment = Alignment(horizontal="center")
        ws_details.cell(row=row_idx, column=9, value=r.get("Error_Details", "None")).alignment = Alignment(horizontal="left")
        
        for c in range(1, 10):
            cell = ws_details.cell(row=row_idx, column=c)
            cell.border = thin_border
            if row_idx % 2 == 1 and c != 6:
                cell.fill = zebra_fill

    # Auto-fit column widths for all sheets
    for sheet in [ws_summary, ws_details]:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if cell.row in [1, 2, 4, 5, 8] and sheet == ws_summary:
                    continue
                if len(val_str) > max_len:
                    max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    wb.save(output_path)
    print(f"[REPORT GENERATED] Successfully saved Excel analysis report to: {output_path}")
    return output_path
